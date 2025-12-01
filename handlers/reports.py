import os
from datetime import datetime
from aiogram import F, Router
from aiogram.fsm.context import FSMContext
from aiogram.types import Message, CallbackQuery, InlineKeyboardButton, FSInputFile, BufferedInputFile
from aiogram.utils.keyboard import InlineKeyboardBuilder
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from states import ReportStates
from utils import get_main_keyboard, edit_or_send
from config import logger

router = Router()
from database import get_database
db = get_database()


@router.callback_query(F.data == "reports_menu")
async def reports_menu(callback: CallbackQuery):
    """Меню отчётов"""
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="👥 База клиентов (Excel)", callback_data="report_clients"))
    builder.row(InlineKeyboardButton(text="💰 Финансовый отчёт (Excel)", callback_data="report_financial"))
    builder.row(InlineKeyboardButton(text="📊 История операций (Excel)", callback_data="report_operations"))
    builder.row(InlineKeyboardButton(text="📦 Отчёт по оборудованию (Excel)", callback_data="report_equipment"))
    builder.row(InlineKeyboardButton(text="◀️ Назад", callback_data="back_to_main"))
    
    await edit_or_send(
        callback,
        "📈 <b>Отчёты и аналитика</b>\n\n"
        "Все отчёты выгружаются в формате Excel\n"
        "Выберите тип отчёта:",
        reply_markup=builder.as_markup(),
        parse_mode='HTML'
    )
    await callback.answer()


@router.callback_query(F.data == "report_clients")
async def report_clients(callback: CallbackQuery, state: FSMContext):
    """Запрос отчёта по клиентам"""
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="❌ Отмена", callback_data="reports_menu"))
    
    await edit_or_send(
        callback,
        "👥 <b>Отчёт по клиентам</b>\n\n"
        "<b>Введите период:</b>\n"
        "ГГГГ-ММ-ДД - ГГГГ-ММ-ДД\n\n"
        "Или отправьте '-' для отчёта за всё время",
        reply_markup=builder.as_markup(),
        parse_mode='HTML'
    )
    await state.update_data(report_type='clients')
    await state.set_state(ReportStates.entering_date_range)
    await callback.answer()


@router.callback_query(F.data == "report_financial")
async def report_financial(callback: CallbackQuery, state: FSMContext):
    """Запрос финансового отчёта"""
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="❌ Отмена", callback_data="reports_menu"))
    
    await edit_or_send(
        callback,
        "💰 <b>Финансовый отчёт</b>\n\n"
        "<b>Введите период:</b>\n"
        "ГГГГ-ММ-ДД - ГГГГ-ММ-ДД\n\n"
        "Или отправьте '-' для отчёта за всё время",
        reply_markup=builder.as_markup(),
        parse_mode='HTML'
    )
    await state.update_data(report_type='financial')
    await state.set_state(ReportStates.entering_date_range)
    await callback.answer()


@router.callback_query(F.data == "report_operations")
async def report_operations(callback: CallbackQuery, state: FSMContext):
    """Запрос отчёта по операциям"""
    builder = InlineKeyboardBuilder()
    builder.row(InlineKeyboardButton(text="❌ Отмена", callback_data="reports_menu"))
    
    await edit_or_send(
        callback,
        "📊 <b>История операций</b>\n\n"
        "<b>Введите период:</b>\n"
        "ГГГГ-ММ-ДД - ГГГГ-ММ-ДД\n\n"
        "Или отправьте '-' для отчёта за всё время",
        reply_markup=builder.as_markup(),
        parse_mode='HTML'
    )
    await state.update_data(report_type='operations')
    await state.set_state(ReportStates.entering_date_range)
    await callback.answer()


@router.callback_query(F.data == "report_equipment")
async def report_equipment(callback: CallbackQuery):
    """Отчёт по загруженности оборудования"""
    await callback.answer("⏳ Формирую отчёт...", show_alert=False)
    
    try:
        filename = generate_equipment_report()
        
        await callback.message.answer_document(
            FSInputFile(filename),
            caption="📦 <b>Отчёт по оборудованию</b>\n\n"
                   "Текущая загруженность и статистика использования",
            parse_mode='HTML'
        )
        
        os.remove(filename)
        logger.info(f"Отчёт по оборудованию отправлен пользователю {callback.from_user.id}")
        
    except Exception as e:
        logger.error(f"Ошибка генерации отчёта по оборудованию: {e}")
        await callback.message.answer(
            "❌ Ошибка при создании отчёта",
            reply_markup=get_main_keyboard()
        )


@router.message(ReportStates.entering_date_range)
async def process_report_dates(message: Message, state: FSMContext):
    """Обработка дат для отчёта"""
    data = await state.get_data()
    report_type = data['report_type']
    
    start_date = None
    end_date = None
    
    if message.text != '-':
        try:
            parts = message.text.split()
            if len(parts) == 3:
                start_date = parts[0].strip()
                end_date = parts[2].strip()
                datetime.strptime(start_date, '%Y-%m-%d')
                datetime.strptime(end_date, '%Y-%m-%d')
            else:
                await message.answer("❌ Неверный формат. Используйте: ГГГГ-ММ-ДД - ГГГГ-ММ-ДД")
                return
        except:
            await message.answer("❌ Неверный формат даты. Проверьте правильность ввода.")
            return
    
    await message.answer("⏳ Формирую отчёт...", reply_markup=get_main_keyboard())
    
    try:
        if report_type == 'clients':
            filename = generate_clients_excel(start_date, end_date)
        elif report_type == 'financial':
            filename = generate_financial_excel(start_date, end_date)
        elif report_type == 'operations':
            filename = generate_operations_excel(start_date, end_date)
        else:
            await message.answer("❌ Неизвестный тип отчёта")
            await state.clear()
            return
        
        period_text = "за всё время"
        if start_date and end_date:
            period_text = f"с {start_date} по {end_date}"
        
        await message.answer_document(
            FSInputFile(filename),
            caption=f"📊 <b>Отчёт готов!</b>\n\nПериод: {period_text}",
            parse_mode='HTML'
        )
        
        os.remove(filename)
        logger.info(f"Отчёт {report_type} отправлен пользователю {message.from_user.id}")
        
    except Exception as e:
        logger.error(f"Ошибка генерации отчёта: {e}", exc_info=True)
        await message.answer(
            "❌ Ошибка при создании отчёта",
            reply_markup=get_main_keyboard()
        )
    
    await state.clear()


def style_header(ws, row_num, columns):
    """Стилизация заголовка таблицы"""
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_num, column_name in enumerate(columns, 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.value = column_name
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border


def style_data_row(ws, row_num, num_columns, is_alt=False):
    """Стилизация строки данных"""
    fill_color = "F2F2F2" if is_alt else "FFFFFF"
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    for col_num in range(1, num_columns + 1):
        cell = ws.cell(row=row_num, column=col_num)
        cell.fill = fill
        cell.border = border
        cell.alignment = Alignment(vertical='center')


def auto_adjust_columns(ws):
    """Автоматическая подстройка ширины столбцов"""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def generate_clients_excel(start_date=None, end_date=None):
    """Генерация Excel отчёта по клиентам"""
    clients = db.get_clients_report(start_date, end_date)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "База клиентов"
    
    # Заголовок отчёта
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = "ОТЧЁТ: БАЗА КЛИЕНТОВ"
    title_cell.font = Font(bold=True, size=14, color="366092")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Период
    ws.merge_cells('A2:F2')
    period_cell = ws['A2']
    if start_date and end_date:
        period_cell.value = f"Период: {start_date} — {end_date}"
    else:
        period_cell.value = "Период: за всё время"
    period_cell.alignment = Alignment(horizontal='center')
    period_cell.font = Font(italic=True)
    
    # Заголовки таблицы
    headers = ['Имя клиента', 'Телефон', 'Первый заказ', 'Последний заказ', 'Всего заказов', 'Общая сумма']
    style_header(ws, 4, headers)
    
    # Данные
    for idx, client in enumerate(clients, start=5):
        client_name, phone, first_order, last_order, total_orders, total_spent = client
        
        ws.cell(row=idx, column=1, value=client_name)
        ws.cell(row=idx, column=2, value=phone)
        ws.cell(row=idx, column=3, value=first_order[:10] if first_order else '')
        ws.cell(row=idx, column=4, value=last_order[:10] if last_order else '')
        ws.cell(row=idx, column=5, value=total_orders)
        ws.cell(row=idx, column=6, value=f"{total_spent:.2f}" if total_spent else "0.00")
        
        style_data_row(ws, idx, 6, is_alt=(idx % 2 == 0))
    
    # Итоги
    summary_row = len(clients) + 6
    ws.merge_cells(f'A{summary_row}:D{summary_row}')
    summary_cell = ws[f'A{summary_row}']
    summary_cell.value = f"ИТОГО КЛИЕНТОВ: {len(clients)}"
    summary_cell.font = Font(bold=True)
    
    total_orders_sum = sum(c[4] for c in clients)
    total_revenue_sum = sum(c[5] if c[5] else 0 for c in clients)
    
    ws.cell(row=summary_row, column=5, value=total_orders_sum).font = Font(bold=True)
    ws.cell(row=summary_row, column=6, value=f"{total_revenue_sum:.2f}").font = Font(bold=True)
    
    auto_adjust_columns(ws)
    
    filename = f"clients_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    return filename


def generate_financial_excel(start_date=None, end_date=None):
    """Генерация Excel финансового отчёта"""
    stats = db.get_financial_report(start_date, end_date)
    orders = db.get_operations_report(start_date, end_date)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Финансовый отчёт"
    
    # Заголовок
    ws.merge_cells('A1:E1')
    title_cell = ws['A1']
    title_cell.value = "ФИНАНСОВЫЙ ОТЧЁТ"
    title_cell.font = Font(bold=True, size=14, color="366092")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Период
    ws.merge_cells('A2:E2')
    period_cell = ws['A2']
    if start_date and end_date:
        period_cell.value = f"Период: {start_date} — {end_date}"
    else:
        period_cell.value = "Период: за всё время"
    period_cell.alignment = Alignment(horizontal='center')
    period_cell.font = Font(italic=True)
    
    # Статистика
    if stats:
        total_orders, total_revenue, avg_order = stats
        
        ws['A4'] = "СВОДКА:"
        ws['A4'].font = Font(bold=True, size=12)
        
        ws['A5'] = "Всего заказов:"
        ws['B5'] = total_orders
        ws['B5'].font = Font(bold=True)
        
        ws['A6'] = "Общая выручка:"
        ws['B6'] = f"{total_revenue:.2f} руб." if total_revenue else "0.00 руб."
        ws['B6'].font = Font(bold=True, color="00AA00")
        
        ws['A7'] = "Средний чек:"
        ws['B7'] = f"{avg_order:.2f} руб." if avg_order else "0.00 руб."
        ws['B7'].font = Font(bold=True)
    
    # Таблица заказов
    ws['A9'] = "ДЕТАЛИЗАЦИЯ ПО ЗАКАЗАМ:"
    ws['A9'].font = Font(bold=True, size=11)
    
    headers = ['№ Заказа', 'Клиент', 'Период', 'Стоимость', 'Статус']
    style_header(ws, 10, headers)
    
    for idx, order in enumerate(orders, start=11):
        order_id, client_name, client_phone, start, end, cost, status, created_at, completed_at = order
        
        ws.cell(row=idx, column=1, value=f"#{order_id}")
        ws.cell(row=idx, column=2, value=f"{client_name} ({client_phone})")
        ws.cell(row=idx, column=3, value=f"{start} — {end}")
        ws.cell(row=idx, column=4, value=cost if cost else "—")
        
        status_map = {
            'pending': 'Ожидает выдачи',
            'issued': 'Выдано',
            'overdue': 'Просрочено',
            'completed': 'Завершено'
        }
        ws.cell(row=idx, column=5, value=status_map.get(status, status))
        
        style_data_row(ws, idx, 5, is_alt=(idx % 2 == 0))
    
    auto_adjust_columns(ws)
    
    filename = f"financial_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    return filename


def generate_operations_excel(start_date=None, end_date=None):
    """Генерация Excel отчёта по операциям"""
    operations = db.get_operations_report(start_date, end_date)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "История операций"
    
    # Заголовок
    ws.merge_cells('A1:H1')
    title_cell = ws['A1']
    title_cell.value = "ИСТОРИЯ ОПЕРАЦИЙ"
    title_cell.font = Font(bold=True, size=14, color="366092")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Период
    ws.merge_cells('A2:H2')
    period_cell = ws['A2']
    if start_date and end_date:
        period_cell.value = f"Период: {start_date} — {end_date}"
    else:
        period_cell.value = "Период: за всё время"
    period_cell.alignment = Alignment(horizontal='center')
    period_cell.font = Font(italic=True)
    
    # Статистика по статусам
    status_stats = {}
    for op in operations:
        status = op[6]
        status_stats[status] = status_stats.get(status, 0) + 1
    
    ws['A4'] = "СТАТИСТИКА:"
    ws['A4'].font = Font(bold=True, size=11)
    
    row = 5
    status_map = {
        'pending': 'Ожидает выдачи',
        'issued': 'Выдано',
        'overdue': 'Просрочено',
        'completed': 'Завершено'
    }
    
    for status, count in status_stats.items():
        ws.cell(row=row, column=1, value=f"{status_map.get(status, status)}:")
        ws.cell(row=row, column=2, value=count)
        ws.cell(row=row, column=2).font = Font(bold=True)
        row += 1
    
    # Таблица операций
    ws[f'A{row + 1}'] = "ДЕТАЛЬНАЯ ИНФОРМАЦИЯ:"
    ws[f'A{row + 1}'].font = Font(bold=True, size=11)
    
    headers = ['№', 'Клиент', 'Телефон', 'Начало', 'Конец', 'Стоимость', 'Статус', 'Создан']
    style_header(ws, row + 2, headers)
    
    data_start_row = row + 3
    for idx, op in enumerate(operations, start=data_start_row):
        order_id, client_name, client_phone, start, end, cost, status, created_at, completed_at = op
        
        ws.cell(row=idx, column=1, value=f"#{order_id}")
        ws.cell(row=idx, column=2, value=client_name)
        ws.cell(row=idx, column=3, value=client_phone)
        ws.cell(row=idx, column=4, value=start)
        ws.cell(row=idx, column=5, value=end)
        ws.cell(row=idx, column=6, value=cost if cost else "—")
        ws.cell(row=idx, column=7, value=status_map.get(status, status))
        ws.cell(row=idx, column=8, value=created_at[:16] if created_at else "")
        
        # Цветовое выделение по статусу
        status_colors = {
            'pending': 'FFF4E6',
            'issued': 'E7F3FF',
            'overdue': 'FFE7E7',
            'completed': 'E7FFE7'
        }
        if status in status_colors:
            for col in range(1, 9):
                ws.cell(row=idx, column=col).fill = PatternFill(
                    start_color=status_colors[status],
                    end_color=status_colors[status],
                    fill_type="solid"
                )
        
        style_data_row(ws, idx, 8, is_alt=False)
    
    auto_adjust_columns(ws)
    
    filename = f"operations_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    return filename


def generate_equipment_report():
    """Генерация отчёта по оборудованию"""
    resources = db.get_resources()
    today = datetime.now().strftime('%Y-%m-%d')
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Оборудование"
    
    # Заголовок
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = "ОТЧЁТ ПО ОБОРУДОВАНИЮ"
    title_cell.font = Font(bold=True, size=14, color="366092")
    title_cell.alignment = Alignment(horizontal='center', vertical='center')
    
    ws.merge_cells('A2:F2')
    date_cell = ws['A2']
    date_cell.value = f"Дата: {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    date_cell.alignment = Alignment(horizontal='center')
    date_cell.font = Font(italic=True)
    
    # Заголовки
    headers = ['Название', 'Всего единиц', 'Доступно сейчас', 'Забронировано', '% загрузки', 'Статус']
    style_header(ws, 4, headers)
    
    for idx, resource in enumerate(resources, start=5):
        res_id, name, description, total_quantity = resource
        
        # Проверяем доступность на сегодня
        available = db.get_available_quantity(res_id, today, today)
        booked = total_quantity - available
        utilization = (booked / total_quantity * 100) if total_quantity > 0 else 0
        
        ws.cell(row=idx, column=1, value=name)
        ws.cell(row=idx, column=2, value=total_quantity)
        ws.cell(row=idx, column=3, value=available)
        ws.cell(row=idx, column=4, value=booked)
        ws.cell(row=idx, column=5, value=f"{utilization:.1f}%")
        
        # Статус
        if available == 0:
            status = "Полностью занято"
            status_color = "FFE7E7"
        elif available < total_quantity * 0.3:
            status = "Высокая загрузка"
            status_color = "FFF4E6"
        else:
            status = "Доступно"
            status_color = "E7FFE7"
        
        ws.cell(row=idx, column=6, value=status)
        
        # Цветовое выделение
        for col in range(1, 7):
            ws.cell(row=idx, column=col).fill = PatternFill(
                start_color=status_color,
                end_color=status_color,
                fill_type="solid"
            )
        
        style_data_row(ws, idx, 6, is_alt=False)
    
    # Итоги
    summary_row = len(resources) + 6
    ws.merge_cells(f'A{summary_row}:F{summary_row}')
    summary_cell = ws[f'A{summary_row}']
    summary_cell.value = f"ВСЕГО ПОЗИЦИЙ: {len(resources)}"
    summary_cell.font = Font(bold=True)
    summary_cell.alignment = Alignment(horizontal='center')
    
    auto_adjust_columns(ws)
    
    filename = f"equipment_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(filename)
    return filename